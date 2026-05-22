"""
ui_backend.py
-------------
Adapter layer between Streamlit UI and backend modules.
UI imports ONLY from this file — never directly from backend.
"""

import asyncio
import os
import io
import json
import tempfile
import threading
from dataclasses import dataclass, asdict
from typing import Callable, Optional

import nest_asyncio
nest_asyncio.apply()


def _run_async(coro):
    """Run a coroutine from any context (main thread or background thread)."""
    try:
        loop = asyncio.get_running_loop()
        return loop.run_until_complete(coro)
    except RuntimeError:
        return asyncio.run(coro)


# ── Config schema ─────────────────────────────────────────────────────────────

@dataclass
class RunConfig:
    api_key:     str   = ""
    api_base:    str   = "https://api.vectorengine.ai/v1"
    model:       str   = "claude-opus-4-7"
    temperature: float = 0.3
    max_tokens:  int   = 2048
    max_workers: int   = 24

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "RunConfig":
        fields = set(RunConfig.__dataclass_fields__)
        return RunConfig(**{k: v for k, v in d.items() if k in fields})


def default_config_from_module() -> RunConfig:
    try:
        import config as _c
        return RunConfig(
            api_key=_c.LLM_API_KEY,
            api_base=_c.LLM_API_BASE,
            model=_c.LLM_MODEL,
            temperature=_c.LLM_TEMPERATURE,
            max_tokens=_c.LLM_MAX_TOKENS,
        )
    except Exception:
        return RunConfig()


# ── Config application ────────────────────────────────────────────────────────

def apply_config(cfg: RunConfig) -> None:
    import config
    from openai import AsyncOpenAI

    config.LLM_API_KEY     = cfg.api_key
    config.LLM_API_BASE    = cfg.api_base
    config.LLM_MODEL       = cfg.model
    config.LLM_TEMPERATURE = cfg.temperature
    config.LLM_MAX_TOKENS  = cfg.max_tokens

    import core.llm_review as llm_review
    old = llm_review._client
    if old is not None:
        try:
            _run_async(old.close())
        except Exception:
            pass
    llm_review._client = AsyncOpenAI(api_key=cfg.api_key, base_url=cfg.api_base)


# ── Processing ────────────────────────────────────────────────────────────────

@dataclass
class RowResult:
    row_num:           int
    error_type:        str
    error_description: str
    source_text:       str
    target_text:       str
    verdict:           Optional[str]
    llm_analysis:      str


def process_file(
    file_bytes: bytes,
    filename: str,
    cfg: RunConfig,
    on_progress: Optional[Callable] = None,
) -> dict:
    """
    Process one Xbench xlsx: parse + LLM review each row.

    on_progress(stage, done, total)
      stage: "parse" | "llm"

    Returns:
      filename, rows (list[RowResult]), stats, error
    """
    apply_config(cfg)

    from core.xbench import parse
    from core.llm_review import llm_secondary_review

    # ── Parse ──────────────────────────────────────────────────────────────
    if on_progress:
        on_progress("parse", 0, 1)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tf.write(file_bytes)
        tmp_path = tf.name
    try:
        rows = parse(tmp_path)
    except Exception as e:
        return {"filename": filename, "rows": [], "stats": {}, "error": f"解析失败：{e}"}
    finally:
        os.unlink(tmp_path)

    if on_progress:
        on_progress("parse", 1, 1)

    valid_rows = [r for r in rows if r["error_description"].strip()]
    if not valid_rows:
        return {"filename": filename, "rows": [], "stats": {"total": 0},
                "error": "无有效数据行"}

    # ── LLM review (async + semaphore) ────────────────────────────────────
    if on_progress:
        on_progress("llm", 0, len(valid_rows))

    sem       = asyncio.Semaphore(cfg.max_workers)
    completed = 0

    async def _review_one(idx, r):
        nonlocal completed
        async with sem:
            try:
                review  = await llm_secondary_review(
                    error_type=r["check_type"],
                    error_description=r["error_description"],
                    source_text=r.get("source_text", ""),
                    target_text=r.get("target_text", ""),
                    search_results=[],
                )
                verdict  = review.final_label
                analysis = review.llm_analysis
            except Exception as e:
                verdict  = "真错误"
                analysis = f"LLM调用失败，默认保留：{e}"

            completed += 1
            if on_progress:
                on_progress("llm", completed, len(valid_rows))

            return idx, RowResult(
                row_num=r["row_num"],
                error_type=r["check_type"],
                error_description=r["error_description"],
                source_text=r.get("source_text", ""),
                target_text=r.get("target_text", ""),
                verdict=verdict,
                llm_analysis=analysis,
            )

    pairs = _run_async(
        asyncio.gather(*[_review_one(i, r) for i, r in enumerate(valid_rows)])
    )
    row_results = [r for _, r in sorted(pairs, key=lambda x: x[0])]

    total      = len(row_results)
    suppressed = sum(1 for r in row_results if r.verdict == "误报")
    kept       = total - suppressed

    return {
        "filename": filename,
        "rows":     row_results,
        "stats":    {"total": total, "suppressed": suppressed, "kept": kept},
    }


# ── Export helpers ────────────────────────────────────────────────────────────

def build_filtered_xlsx(file_bytes: bytes, row_results: list[RowResult]) -> bytes:
    """
    Return a filtered Xbench xlsx with 误报 rows deleted.
    Uses xbench.filter_xlsx internally.
    """
    from core.xbench import filter_xlsx

    keep_rows = {r.row_num for r in row_results if r.verdict == "真错误"}

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tf.write(file_bytes)
        src_path = tf.name

    dst_fd, dst_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(dst_fd)
    try:
        filter_xlsx(src_path, keep_rows, dst_path)
        with open(dst_path, "rb") as f:
            return f.read()
    finally:
        os.unlink(src_path)
        if os.path.exists(dst_path):
            os.unlink(dst_path)


def build_analysis_xlsx(row_results: list[RowResult]) -> bytes:
    """Return analysis xlsx showing verdict + reasoning for every row."""
    import openpyxl
    from openpyxl.styles import PatternFill

    red_fill   = PatternFill("solid", fgColor="FFCCCC")
    green_fill = PatternFill("solid", fgColor="CCFFCC")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "复核结果"
    ws.append(["#", "原文", "译文", "错误类型", "错误描述", "系统判断", "复核依据"])

    for i, r in enumerate(row_results, 1):
        ws.append([
            i,
            r.source_text, r.target_text,
            r.error_type,  r.error_description,
            r.verdict,     r.llm_analysis,
        ])
        fill = green_fill if r.verdict == "误报" else red_fill
        for col in range(1, 8):
            ws.cell(i + 1, col).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_zip_bytes(results: list[dict], file_bytes_map: dict) -> bytes:
    """
    Pack per-file outputs into a zip.
    file_bytes_map: {filename: original_bytes}
    """
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            if not r.get("rows") or r.get("error"):
                continue
            stem = os.path.splitext(r["filename"])[0]
            orig = file_bytes_map.get(r["filename"])
            if orig:
                filtered = build_filtered_xlsx(orig, r["rows"])
                zf.writestr(f"{stem}_filtered.xlsx", filtered)
            analysis = build_analysis_xlsx(r["rows"])
            zf.writestr(f"{stem}_analysis.xlsx", analysis)
    return buf.getvalue()


# ── Model management ─────────────────────────────────────────────────────────

PRESET_MODELS = [
    "claude-opus-4-7",
    "claude-sonnet-4-6",
    "claude-haiku-4-5-20251001",
]

_CUSTOM_MODELS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "custom_models.json"
)


def _load_model_data() -> dict:
    if not os.path.exists(_CUSTOM_MODELS_FILE):
        return {"custom": [], "disabled_presets": []}
    with open(_CUSTOM_MODELS_FILE, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return {"custom": data, "disabled_presets": []}
    return data


def _save_model_data(data: dict) -> None:
    os.makedirs(os.path.dirname(_CUSTOM_MODELS_FILE), exist_ok=True)
    with open(_CUSTOM_MODELS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


_FALLBACK_MODEL = "gemini-3.1-flash-lite"


def get_all_models() -> list[str]:
    data     = _load_model_data()
    disabled = set(data.get("disabled_presets", []))
    custom   = [m for m in data.get("custom", []) if m not in PRESET_MODELS]
    models   = [m for m in PRESET_MODELS if m not in disabled] + custom
    if not models:
        add_custom_model(_FALLBACK_MODEL)
        models = [_FALLBACK_MODEL]
    return models


def add_custom_model(name: str) -> None:
    data = _load_model_data()
    if name in PRESET_MODELS:
        disabled = data.get("disabled_presets", [])
        if name in disabled:
            data["disabled_presets"] = [m for m in disabled if m != name]
            _save_model_data(data)
        return
    if name not in data["custom"]:
        data["custom"].append(name)
        _save_model_data(data)


def remove_model(name: str) -> None:
    data = _load_model_data()
    if name in data.get("custom", []):
        data["custom"] = [m for m in data["custom"] if m != name]
    elif name in PRESET_MODELS:
        data.setdefault("disabled_presets", [])
        if name not in data["disabled_presets"]:
            data["disabled_presets"].append(name)
    _save_model_data(data)


# ── Background processing task ───────────────────────────────────────────────

class ProcessingTask:
    def __init__(self, file_names: list[str]):
        self.total        = len(file_names)
        self.file_names   = file_names
        self.current      = 0
        self.stage        = ""
        self.stage_done   = 0
        self.stage_total  = 0
        self.results: list[dict] = []
        self._done_event  = threading.Event()

    @property
    def done(self) -> bool:
        return self._done_event.is_set()

    @property
    def current_name(self) -> str:
        return self.file_names[self.current] if self.current < self.total else ""


def start_processing_task(
    files: list[tuple[str, bytes]],
    cfg: RunConfig,
) -> ProcessingTask:
    task = ProcessingTask([name for name, _ in files])

    def _worker():
        for i, (name, data) in enumerate(files):
            task.current     = i
            task.stage       = ""
            task.stage_done  = 0
            task.stage_total = 0

            def on_progress(stage, done, total):
                task.stage       = stage
                task.stage_done  = done
                task.stage_total = total

            result = process_file(data, name, cfg, on_progress)
            task.results.append(result)

        task._done_event.set()

    threading.Thread(target=_worker, daemon=True).start()
    return task


# ── API test ──────────────────────────────────────────────────────────────────

def test_api_connection(cfg: RunConfig) -> tuple[bool, str]:
    """Send a minimal request to verify API key + base URL. Returns (ok, message)."""
    from openai import OpenAI, AuthenticationError, APIConnectionError
    try:
        client = OpenAI(api_key=cfg.api_key, base_url=cfg.api_base)
        resp   = client.chat.completions.create(
            model=cfg.model,
            messages=[{"role": "user", "content": "hi"}],
            max_tokens=5,
            temperature=0,
        )
        return True, f"连接成功 · 模型：{resp.model}"
    except AuthenticationError:
        return False, "认证失败：API Key 无效"
    except APIConnectionError as e:
        return False, f"连接失败：{e}"
    except Exception as e:
        return False, f"错误：{e}"
